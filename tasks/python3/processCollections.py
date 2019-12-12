#!/usr/bin/env python

from concurrent.futures import ProcessPoolExecutor, ThreadPoolExecutor
from optparse import OptionParser
from openpyxl import load_workbook
from pprint import pprint as pp
import os
import json
import sys

prog = os.path.basename(__file__)
product_dict = {}
wv_product_dict = {}

parser = OptionParser(usage="Usage: %s <input_dir> <output_file>" % prog)
(options, args) = parser.parse_args()
input_dir = args[0]
output_file = args[1]

# Validate that the concept_id value is not a placeholder and handle multiple values
def validate_add_concept_id(wv_id, concept_id):
  # TODO need to remove spaces from concept ids or do URL encoding
  if concept_id.startswith("TBD") or concept_id.startswith("N/A"):
    return
  if concept_id is not None:
    split_ids = concept_id.split()
    # Some values have spaces, sometimes spaces represent separate IDs
    # Here we are assuming if it starts with a C, each value is a separate ID
    if (len(split_ids) > 1) and all(c_id.startswith('C') for c_id in split_ids):
      # TODO re-enable or otherwise handle multiple concept ids
      # wv_product_dict[wv_id] = split_ids
      print('%s WARNING: multiple ids for product %s' % (prog, wv_id))
    else:
      wv_product_dict[wv_id] = concept_id


# Get the product ids and concept ids from the respective sheets and add to product_dict
def get_values_from_sheets(prod_id_sheet, prod_metadata_sheet):
  # TODO should not set arbitrary max number of rows but instead
  # read as many non-empty rows as we can find
  for row in prod_id_sheet.iter_rows(min_row=3, max_col=3, max_row=200):
    prod_id = row[0].value
    wv_id = row[2].value
    if prod_id is not None and wv_id is not None:
      product_dict[prod_id] =  wv_id

  for md_row in prod_metadata_sheet.iter_rows(min_row=3, max_col=2, max_row=200):
    prod_id = md_row[0].value
    concept_id = md_row[1].value
    if prod_id is not None and concept_id is not None:
      wv_id = product_dict[prod_id]
      validate_add_concept_id(wv_id, concept_id)

def process_sheet(filePath):
  filename = os.path.basename(filePath)
  try:
    workbook = load_workbook(filePath)
    prod_id_sheet = workbook['Product Identification']
    prod_metadata_sheet = workbook['Product Metadata']
    get_values_from_sheets(prod_id_sheet, prod_metadata_sheet)
  except Exception as e:
    sys.stderr.write("%s ERROR: [%s]: %s\n" % (prog, filename, str(e)))
    # TODO should exit on error once these are fixed
    # sys.exit(1)

#MAIN
with ThreadPoolExecutor() as executor:
  for root, dirs, files in os.walk(input_dir):
    for file in files:
      fp = os.path.join(input_dir, file)
      executor.submit(process_sheet, fp)

with open(output_file, "w") as ofp:
  json.dump(wv_product_dict, ofp)

print("%s: Mapped %s collections to products in %s" % (
  prog,
  len(wv_product_dict),
  os.path.basename(output_file)
))